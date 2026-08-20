"""
Process returned annotation workbooks.

Computes:
  1. Fleiss' kappa among the three NEW annotators (comparable to the 0.71 the
     organisers reported)
  2. Disagreement rate vs. the published gold label, per stratum
  3. Stratum-weighted extrapolation to the full 10,200-item test set, with CIs
  4. Distribution of the C1-C4 error categories
  5. Confusion matrix: published gold vs. adjudicated new label

Usage:  python 05_analyse_annotations.py --tag pilot
        python 05_analyse_annotations.py --tag pilot --simulate   (self-test)
"""
import pandas as pd, numpy as np, json, argparse
from pathlib import Path
from collections import Counter
from openpyxl import load_workbook

BASE = Path("/home/claude/audit")
ANN = BASE / "annotation"; RES = BASE / "results"; RES.mkdir(exist_ok=True)

ap = argparse.ArgumentParser()
ap.add_argument("--tag", default="pilot")
ap.add_argument("--simulate", action="store_true",
                help="generate synthetic annotations to test the pipeline")
args = ap.parse_args()

gold = pd.read_csv(ANN / f"audit_sample_{args.tag}_WITH_GOLD.csv", dtype=str,
                   keep_default_na=False, na_values=[])
gold["item_no"] = gold["item_no"].astype(int)
TYPES = ["None", "Abusive", "Sexism", "Religious Hate", "Political Hate", "Profane"]
ERRS = ["C0-agree", "C1-label mismatch", "C2-needs context",
        "C3-schema gap", "C4-ambiguous/ironic"]

# ---------------- load or simulate ----------------
def read_wb(path):
    wb = load_workbook(path, data_only=True)
    ws = [s for s in wb.sheetnames if s.startswith("annotate")][0]
    ws = wb[ws]
    rows = []
    for r in range(3, ws.max_row + 1):
        item = ws.cell(r, 1).value
        if item is None:
            continue
        rows.append(dict(item_no=int(item), hate_type=ws.cell(r, 4).value,
                         target=ws.cell(r, 5).value, severity=ws.cell(r, 6).value,
                         error_category=ws.cell(r, 7).value,
                         confidence=ws.cell(r, 8).value, notes=ws.cell(r, 9).value))
    return pd.DataFrame(rows)

anns = {}
if args.simulate:
    print(">>> SIMULATION MODE — synthetic annotations, for pipeline testing only.\n")
    rng = np.random.default_rng(7)
    # Annotators agree with gold more often on the control stratum than on the
    # diagnostic strata; this is the pattern we expect, not a claim about reality.
    p_agree = {"S1_no_target": .55, "S2_sev_tension": .60, "S3_rare_classes": .70,
               "S4_confusable": .75, "S5_none_control": .93}
    for a in ["A1", "A2", "A3"]:
        rows = []
        for _, r in gold.iterrows():
            pa = p_agree[r.stratum]
            if rng.random() < pa:
                t, e = r.hate_type, "C0-agree"
            else:
                t = rng.choice([x for x in TYPES if x != r.hate_type])
                e = rng.choice(ERRS[1:], p=[.45, .25, .15, .15])
            rows.append(dict(item_no=int(r.item_no), hate_type=t, target=r.to_whom,
                             severity=r.hate_severity, error_category=e,
                             confidence=rng.choice(["High", "Medium", "Low"], p=[.5, .35, .15]),
                             notes=""))
        anns[a] = pd.DataFrame(rows)
else:
    for a in ["A1", "A2", "A3"]:
        p = ANN / f"annotation_{args.tag}_{a}.xlsx"
        df = read_wb(p)
        df = df[df.hate_type.notna()]
        if df.empty:
            raise SystemExit(f"{p.name} has no completed rows yet.")
        anns[a] = df
        print(f"   loaded {a}: {len(df)} completed items")

# ---------------- 1. Fleiss' kappa among new annotators ----------------
def fleiss_kappa(mat):
    """mat: n_items x n_categories counts."""
    n_items, _ = mat.shape
    n_rat = mat.sum(axis=1)
    if not (n_rat == n_rat[0]).all():
        raise ValueError("unequal raters per item")
    n = n_rat[0]
    P_i = ((mat ** 2).sum(axis=1) - n) / (n * (n - 1))
    P_bar = P_i.mean()
    p_j = mat.sum(axis=0) / (n_items * n)
    P_e = (p_j ** 2).sum()
    return (P_bar - P_e) / (1 - P_e), P_bar, P_e

merged = gold[["item_no", "stratum", "hate_type", "to_whom", "hate_severity"]].rename(
    columns={"hate_type": "gold_type"})
for a, df in anns.items():
    merged = merged.merge(df[["item_no", "hate_type", "error_category", "confidence"]]
                          .rename(columns={"hate_type": f"{a}_type",
                                           "error_category": f"{a}_err",
                                           "confidence": f"{a}_conf"}),
                          on="item_no", how="inner")
print(f"\n   items with all three annotations: {len(merged)}")

mat = np.zeros((len(merged), len(TYPES)), dtype=int)
for i, (_, r) in enumerate(merged.iterrows()):
    for a in anns:
        v = r[f"{a}_type"]
        if v in TYPES:
            mat[i, TYPES.index(v)] += 1
keep = mat.sum(axis=1) == 3
kappa, Pbar, Pe = fleiss_kappa(mat[keep])

print("\n" + "=" * 70)
print("1. INTER-ANNOTATOR AGREEMENT (new annotators, hate type)")
print("=" * 70)
print(f"   Fleiss' kappa       = {kappa:.4f}   (organisers reported 0.71)")
print(f"   observed agreement  = {Pbar:.4f}")
print(f"   chance agreement    = {Pe:.4f}")
verdict = ("REPLICATES the reported reliability" if kappa >= 0.65 else
           "FAILS TO REPLICATE — our annotators agree substantially less")
print(f"   -> {verdict}")

# ---------------- 2. majority label & disagreement with gold ----------------
def majority(row):
    votes = [row[f"{a}_type"] for a in anns if row[f"{a}_type"] in TYPES]
    if not votes:
        return None, 0
    c = Counter(votes).most_common()
    return c[0][0], c[0][1]

merged[["new_type", "n_votes"]] = merged.apply(
    lambda r: pd.Series(majority(r)), axis=1)
merged["unanimous"] = merged.n_votes == 3
merged["has_majority"] = merged.n_votes >= 2
# PRIMARY estimate: >=2 of 3 annotators agree on a label different from gold.
merged["majority_disagreement"] = merged.has_majority & (merged.new_type != merged.gold_type)
# CONSERVATIVE lower bound: all three independently converge on the same
# alternative label. Very strict - three annotators rarely converge on the same
# wrong label by chance, so this under-counts real errors.
merged["confident_disagreement"] = merged.unanimous & (merged.new_type != merged.gold_type)
# Items where all three annotators disagree with each other carry no majority
# and are reported separately: they are evidence of an ill-posed item, not of a
# specific mislabelling.
merged["no_majority"] = ~merged.has_majority
merged["disagrees_gold"] = merged.majority_disagreement

print("\n" + "=" * 70)
print("2. DISAGREEMENT WITH THE PUBLISHED GOLD LABEL, BY STRATUM")
print("=" * 70)
POOLS = {"S1_no_target": 342, "S2_sev_tension": 843, "S3_rare_classes": 208,
         "S4_confusable": 3108, "S5_none_control": 5751}
N_TEST = 10200

def wilson(k, n, z=1.96):
    if n == 0: return (0., 0.)
    p = k / n
    d = 1 + z**2/n; c = p + z**2/(2*n)
    h = z*np.sqrt(p*(1-p)/n + z**2/(4*n**2))
    return ((c-h)/d, (c+h)/d)

print(f"\n   {'stratum':<18}{'n':>4}{'maj.disag':>11}{'rate':>8}{'95% CI':>18}"
      f"{'unanim':>8}{'no-maj':>8}")
strat_rows = []
for s, g in merged.groupby("stratum"):
    n = len(g)
    k = int(g.majority_disagreement.sum())
    kc = int(g.confident_disagreement.sum())
    nm = int(g.no_majority.sum())
    lo, hi = wilson(k, n)
    print(f"   {s:<18}{n:>4}{k:>11}{k/n:>8.3f}   [{lo:.3f}, {hi:.3f}]"
          f"{kc:>8}{nm:>8}")
    strat_rows.append(dict(stratum=s, n=n, majority_disagree=k,
                           maj_rate=round(k/n, 4), ci=[round(lo, 4), round(hi, 4)],
                           unanimous_disagree=kc, unan_rate=round(kc/n, 4),
                           no_majority=nm, pool=POOLS[s]))

# ---------------- 3. stratum-weighted extrapolation ----------------
print("\n" + "=" * 70)
print("3. EXTRAPOLATION TO THE FULL TEST SET (n = 10,200)")
print("=" * 70)
def extrapolate(key):
    est, var = 0.0, 0.0
    for r in strat_rows:
        w = r["pool"] / N_TEST
        p = r[key]
        est += w * p
        var += (w ** 2) * p * (1 - p) / max(r["n"], 1)
    return est, np.sqrt(var)

est, se = extrapolate("maj_rate")
est_c, se_c = extrapolate("unan_rate")
print(f"\n   PRIMARY  (majority of 3 disagrees with gold)")
print(f"      rate {est:.4f}  95% CI [{max(0,est-1.96*se):.4f}, {min(1,est+1.96*se):.4f}]")
print(f"      implied mislabelled items: ~{est*N_TEST:.0f} "
      f"[{max(0,est-1.96*se)*N_TEST:.0f}, {min(1,est+1.96*se)*N_TEST:.0f}]")
print(f"\n   CONSERVATIVE  (all 3 converge on the same alternative label)")
print(f"      rate {est_c:.4f}  95% CI [{max(0,est_c-1.96*se_c):.4f}, "
      f"{min(1,est_c+1.96*se_c):.4f}]")
print(f"      implied mislabelled items: ~{est_c*N_TEST:.0f}")
print("\n   Report BOTH. The primary rate is the headline; the conservative rate")
print("   is the floor that survives the strictest possible reading.")
print("\n   For reference, Jin et al. (CIDR 2026) report 52.8% (BIRD Mini-Dev)")
print("   and 66.1% (Spider 2.0-Snow) on text-to-SQL benchmarks.")

# ---------------- 4. error categories ----------------
print("\n" + "=" * 70)
print("4. ERROR-CATEGORY DISTRIBUTION (all annotator judgements pooled)")
print("=" * 70)
allerr = pd.concat([merged[f"{a}_err"] for a in anns])
vc = allerr.value_counts()
print()
for k, v in vc.items():
    print(f"   {str(k):<24}{v:>6}  {100*v/len(allerr):>5.1f}%")

# ---------------- 5. confusion ----------------
print("\n" + "=" * 70)
print("5. CONFUSION — published gold (rows) vs. adjudicated new label (cols)")
print("=" * 70)
cm = pd.crosstab(merged.gold_type, merged.new_type)
print()
print(cm.to_string())

out = dict(tag=args.tag, simulated=args.simulate, n_items=len(merged),
           fleiss_kappa=round(float(kappa), 4), reported_kappa=0.71,
           observed_agreement=round(float(Pbar), 4),
           strata=strat_rows,
           extrapolated_error_rate=round(float(est), 4),
           extrapolated_conservative=round(float(est_c), 4),
           extrapolated_ci=[round(float(max(0, est-1.96*se)), 4),
                            round(float(min(1, est+1.96*se)), 4)],
           implied_mislabelled=int(est*N_TEST),
           error_categories={str(k): int(v) for k, v in vc.items()})
with open(RES / f"annotation_analysis_{args.tag}"
          f"{'_SIM' if args.simulate else ''}.json", "w") as f:
    json.dump(out, f, indent=2, ensure_ascii=False)
merged.to_csv(RES / f"merged_annotations_{args.tag}"
              f"{'_SIM' if args.simulate else ''}.csv", index=False)
print(f"\n\n-> results/annotation_analysis_{args.tag}"
      f"{'_SIM' if args.simulate else ''}.json")
if args.simulate:
    print("\n*** SIMULATED NUMBERS. Pipeline verified; they mean nothing empirically. ***")
