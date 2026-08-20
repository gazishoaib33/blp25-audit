"""
Build the stratified audit sample and a blind annotation workbook.

Strata are chosen from the findings of scripts 01-03:
  S1 coherence_no_target : hate_type != None BUT to_whom == None   (342 items)
  S2 severity_tension    : hate_type != None BUT severity == Little to None (986)
  S3 rare_classes        : Sexism (29) + Religious Hate (179)
  S4 confusable_head     : Abusive / Profane / Political Hate
  S5 none_control        : hate_type == None  (control stratum)

Annotators are BLIND: gold labels live only in the KEY sheet, which is
written to a separate file that annotators never receive.
"""
import pandas as pd, numpy as np, json, argparse
from pathlib import Path
import pathlib
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

import os
_HERE = pathlib.Path(__file__).resolve().parent
_ROOT = _HERE.parent
# Dataset location: env var BLP25_DATA, else ../blp25_task1, else ./blp25_task1
_CANDIDATES = [os.environ.get("BLP25_DATA"), _ROOT / "blp25_task1",
               _ROOT.parent / "blp25_task1"]
_BASE = next((pathlib.Path(c) for c in _CANDIDATES if c and pathlib.Path(c).exists()), None)
if _BASE is None:
    raise SystemExit("Dataset not found. Clone https://github.com/AridHasan/blp25_task1 "
                     "next to this repo, or set BLP25_DATA=/path/to/blp25_task1")

DATA = _BASE / "data"
OUT = _ROOT / "annotation"; OUT.mkdir(parents=True, exist_ok=True)
SEED = 20260819

ap = argparse.ArgumentParser()
ap.add_argument("--n", type=int, default=500, help="total sample size")
ap.add_argument("--pilot", action="store_true", help="build a 120-item pilot instead")
args = ap.parse_args()
TOTAL = 120 if args.pilot else args.n
TAG = "pilot" if args.pilot else "full"

rng = np.random.default_rng(SEED)
d = pd.read_csv(DATA / "subtask_1C" / "blp25_hatespeech_subtask_1C_test_with_labels.tsv",
                sep="\t", dtype=str, keep_default_na=False, na_values=[])

# ---------------- define strata ----------------
nz = d.hate_type != "None"
strata = {
    "S1_no_target":      d[nz & (d.to_whom == "None")],
    "S2_sev_tension":    d[nz & (d.hate_severity == "Little to None") & (d.to_whom != "None")],
    "S3_rare_classes":   d[d.hate_type.isin(["Sexism", "Religious Hate"])],
    "S4_confusable":     d[d.hate_type.isin(["Abusive", "Profane", "Political Hate"])
                           & (d.to_whom != "None") & (d.hate_severity != "Little to None")],
    "S5_none_control":   d[d.hate_type == "None"],
}
# allocation: over-sample the diagnostic strata, keep a real control
alloc = {"S1_no_target": .24, "S2_sev_tension": .20, "S3_rare_classes": .20,
         "S4_confusable": .20, "S5_none_control": .16}

print("=" * 70)
print(f"BUILDING {TAG.upper()} AUDIT SAMPLE  (target n = {TOTAL})")
print("=" * 70)
picked = []
for name, frame in strata.items():
    want = int(round(TOTAL * alloc[name]))
    take = min(want, len(frame))
    if name == "S3_rare_classes":
        # take ALL sexism items (only 29 exist), fill the rest with religious hate
        sx = frame[frame.hate_type == "Sexism"]
        rh = frame[frame.hate_type == "Religious Hate"]
        n_rh = max(0, take - len(sx))
        sel = pd.concat([sx, rh.sample(min(n_rh, len(rh)), random_state=SEED)])
    else:
        sel = frame.sample(take, random_state=SEED)
    sel = sel.assign(stratum=name)
    picked.append(sel)
    print(f"   {name:<20} pool={len(frame):>6}  sampled={len(sel):>4}")

sample = pd.concat(picked, ignore_index=True).drop_duplicates(subset=["id"])
sample = sample.sample(frac=1.0, random_state=SEED).reset_index(drop=True)  # shuffle
sample.insert(0, "item_no", range(1, len(sample) + 1))
print(f"\n   TOTAL sampled (deduped): {len(sample)}")

sample.to_csv(OUT / f"audit_sample_{TAG}_WITH_GOLD.csv", index=False)
print(f"   key file -> annotation/audit_sample_{TAG}_WITH_GOLD.csv  (DO NOT SHARE)")

# ---------------- workbook ----------------
HEAD_FILL = PatternFill("solid", fgColor="1F3864")
HEAD_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")
BODY = Font(name="Arial", size=10)
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
EX_FILL = PatternFill("solid", fgColor="E2EFDA")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

TYPES = ["None", "Abusive", "Sexism", "Religious Hate", "Political Hate", "Profane"]
TARGETS = ["None", "Individual", "Organization", "Community", "Society"]
SEVS = ["Little to None", "Mild", "Severe"]
ERRS = ["C0-agree", "C1-label mismatch", "C2-needs context",
        "C3-schema gap", "C4-ambiguous/ironic"]
CONF = ["High", "Medium", "Low"]

def build_sheet(ws, df, annot):
    ws.sheet_view.showGridLines = False
    cols = ["item_no", "id", "text", "your_hate_type", "your_target",
            "your_severity", "error_category", "confidence", "notes"]
    widths = [8, 10, 78, 17, 15, 16, 22, 12, 40]
    for j, (c, w) in enumerate(zip(cols, widths), start=1):
        cell = ws.cell(row=1, column=j, value=c)
        cell.fill, cell.font = HEAD_FILL, HEAD_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = "D2"
    ws.row_dimensions[1].height = 30

    # example row (row 2), clearly marked
    ex = ["EXAMPLE", "—", "এই লোকটা একটা চোর, ওকে ধরে পুলিশে দেওয়া উচিত",
          "Abusive", "Individual", "Mild", "C0-agree", "High",
          "Gold said Abusive/Individual/Mild — I agree. / গোল্ড লেবেলের সাথে একমত।"]
    for j, v in enumerate(ex, start=1):
        c = ws.cell(row=2, column=j, value=v)
        c.font = Font(name="Arial", size=10, italic=True)
        c.fill = EX_FILL; c.border = BORDER
        c.alignment = Alignment(vertical="top", wrap_text=(j in (3, 9)))

    start = 3
    for i, r in df.iterrows():
        row = start + i
        ws.cell(row=row, column=1, value=int(r.item_no)).font = BODY
        ws.cell(row=row, column=2, value=str(r.id)).font = BODY
        tc = ws.cell(row=row, column=3, value=r.text)
        tc.font = BODY; tc.alignment = Alignment(vertical="top", wrap_text=True)
        for j in range(4, 10):
            c = ws.cell(row=row, column=j)
            c.fill = INPUT_FILL; c.font = BODY; c.border = BORDER
            c.alignment = Alignment(vertical="top", wrap_text=(j == 9))
        ws.row_dimensions[row].height = 34

    end = start + len(df) - 1
    for col, opts in [("D", TYPES), ("E", TARGETS), ("F", SEVS),
                      ("G", ERRS), ("H", CONF)]:
        dv = DataValidation(type="list", formula1='"' + ",".join(opts) + '"',
                            allow_blank=True, showDropDown=False)
        dv.error = "Pick a value from the dropdown."
        ws.add_data_validation(dv)
        dv.add(f"{col}{start}:{col}{end}")
    return end

def build_instructions(ws, n_items):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 96
    rows = [
        ("BLP-2025 Task 1 — Annotation Audit", ""),
        ("", ""),
        ("What you are doing", "Re-labelling a sample of comments from the BLP-2025 Bangla hate-speech "
                               "test set, WITHOUT seeing the original label. We compare your labels to "
                               "the published ones afterwards to estimate the dataset's error rate."),
        ("আপনি কী করছেন", "BLP-2025 বাংলা হেট-স্পিচ টেস্ট সেটের কিছু মন্তব্য আপনি নতুন করে লেবেল করবেন। "
                          "মূল লেবেল আপনি দেখতে পাবেন না — এটাই উদ্দেশ্য।"),
        ("", ""),
        ("CRITICAL", "Do NOT look up the original labels. Do NOT discuss items with the other "
                     "annotators until all three of you have finished. Blindness is the entire point."),
        ("", ""),
        ("Items to label", n_items),
        ("Expected time", f"about {round(n_items*0.75/60,1)}–{round(n_items*1.5/60,1)} hours"),
        ("Yellow cells", "the only cells you edit. Row 2 is a filled-in EXAMPLE — do not overwrite it."),
        ("", ""),
        ("STEP 1 — label it yourself", "Read the comment. Choose hate_type, target, severity using your own "
                                        "judgement and the definitions below."),
        ("STEP 2 — flag the difficulty", "Choose the error_category that describes why this item is hard "
                                          "(or C0 if it is not hard)."),
        ("STEP 3 — confidence", "High / Medium / Low. Be honest — Low is useful data, not failure."),
        ("STEP 4 — notes", "One line, English or Bangla, whenever you pick C1–C4."),
        ("", ""),
        ("=== HATE TYPE ===", ""),
        ("None", "No hate, abuse, or profanity. / কোনো বিদ্বেষ বা গালি নেই।"),
        ("Abusive", "Insulting, demeaning, or degrading a person or group. / অপমানজনক বা হেয় করা।"),
        ("Sexism", "Hostility based on gender or sexuality. / লিঙ্গভিত্তিক বিদ্বেষ।"),
        ("Religious Hate", "Hostility toward a religion or its followers. / ধর্মভিত্তিক বিদ্বেষ।"),
        ("Political Hate", "Hostility toward a political party, leader, or ideology. / রাজনৈতিক বিদ্বেষ।"),
        ("Profane", "Obscene or vulgar language, not necessarily targeted. / অশ্লীল ভাষা।"),
        ("", ""),
        ("=== TARGET ===", ""),
        ("None", "No identifiable target."),
        ("Individual", "A specific named or clearly implied person."),
        ("Organization", "A party, company, institution, or media outlet."),
        ("Community", "A religious, ethnic, gender, or regional group."),
        ("Society", "Society at large, or the country as a whole."),
        ("", ""),
        ("=== SEVERITY ===", ""),
        ("Little to None", "Minimal or no hateful intent."),
        ("Mild", "Moderately offensive, or implicitly hateful."),
        ("Severe", "Strongly derogatory, or inciting harm."),
        ("", ""),
        ("=== ERROR CATEGORY (adapted from Jin et al., CIDR 2026) ===", ""),
        ("C0-agree", "Straightforward. One reading, and you are confident in it."),
        ("C1-label mismatch", "The item is clear, but you expect the published label got it wrong — e.g. "
                              "criticism of a government read as hate, or a quoted slur read as the "
                              "speaker's own. / মন্তব্য স্পষ্ট, কিন্তু লেবেল ভুল মনে হচ্ছে।"),
        ("C2-needs context", "Unjudgeable without the parent video, the thread, or the news event. These "
                             "are YouTube comments stripped of context. / প্রসঙ্গ ছাড়া বোঝা যাচ্ছে না।"),
        ("C3-schema gap", "Genuinely hateful but no category fits, OR two categories fit equally well "
                          "(religious + political is the common case). / স্কিমা মানানসই নয়।"),
        ("C4-ambiguous/ironic", "Sarcasm, irony, reclaimed slurs, or reported speech. The literal and "
                                "intended readings differ. / ব্যঙ্গ বা দ্ব্যর্থবোধক।"),
        ("", ""),
        ("Rules of thumb", "Judge the COMMENT, not the commenter, and not the topic. Criticism of a "
                           "policy or a public figure's actions is not automatically hate. Profanity "
                           "aimed at nobody is Profane with target None. When two types fit, pick the "
                           "more specific one and flag C3."),
    ]
    for i, (a, b) in enumerate(rows, start=1):
        ca, cb = ws.cell(row=i, column=1, value=a), ws.cell(row=i, column=2, value=b)
        bold = a.startswith("===") or a.isupper() or a.startswith("STEP") or i == 1
        ca.font = Font(name="Arial", size=12 if i == 1 else 10, bold=bold or i == 1,
                       color="1F3864" if a.startswith("===") or i == 1 else "000000")
        cb.font = Font(name="Arial", size=10)
        cb.alignment = Alignment(wrap_text=True, vertical="top")
        ca.alignment = Alignment(wrap_text=True, vertical="top")
        if a == "CRITICAL":
            cb.font = Font(name="Arial", size=10, bold=True, color="C00000")

blind = sample[["item_no", "id", "text"]].reset_index(drop=True)
for annot in ["A1", "A2", "A3"]:
    wb = Workbook()
    ws0 = wb.active; ws0.title = "Instructions"
    build_instructions(ws0, len(blind))
    ws = wb.create_sheet(f"annotate_{annot}")
    build_sheet(ws, blind, annot)
    path = OUT / f"annotation_{TAG}_{annot}.xlsx"
    wb.save(path)
    print(f"   workbook -> annotation/{path.name}")

print("\nDone. Send one workbook to each annotator; keep the *_WITH_GOLD.csv yourself.")
